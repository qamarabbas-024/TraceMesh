"""Excel XLSX investigation workbook exporter."""
import io


def export_to_xlsx(data: dict) -> bytes:
    """Export investigation findings into a multi-tab or structured Excel XLSX sheet."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OSINT Findings"

        # Headers
        ws.append(["Service Module", "Attribute", "Discovered Value"])
        header_fill = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
        header_font = Font(color="00F0FF", bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        for module, result in data.items():
            if module == "_meta":
                continue
            if isinstance(result, dict):
                for k, v in result.items():
                    ws.append([module, k, str(v)[:500]])
            else:
                ws.append([module, "Result", str(result)[:500]])

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    except Exception as e:
        return b""
